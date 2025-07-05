import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
import os

"""Kiểm tra file đã tồn tại chưa"""
def existfile(file_path):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        # sheet = workbook.active
        # sheet.title = "Biểu đồ"
        workbook.save(file_path)
    return openpyxl .load_workbook(file_path)
    
def save_bar_chart(visitor_data, workbook):
    """Tạo và lưu biểu đồ cột theo kênh lưu lượng."""
    channel_method_by_id = visitor_data.groupby("Traffic channel")["Visitor id"].count()
    plt.figure(figsize=(7, 4))
    plt.bar(channel_method_by_id.index, channel_method_by_id.values)
    plt.savefig("Channel_by_visitor_id.png")
    plt.close()  
    img = Image("Channel_by_visitor_id.png")
    bar_sheet = workbook.create_sheet(title="Biểu đồ Cột")
    bar_sheet.add_image(img, 'A1')
    print("Đã thêm thành công")

def save_pie_chart(visitor_data, workbook):
    """Tạo và lưu biểu đồ tròn theo loại thiết bị."""
    device_by_id = visitor_data.groupby("Device type")["Visitor id"].count()
    plt.figure(figsize=(7, 4))
    plt.pie(device_by_id, labels=device_by_id.index, autopct="%1.2f%%")
    plt.savefig("Device_by_id.png")
    plt.close() 
    img2 = Image("Device_by_id.png")
    pie_sheet = workbook.create_sheet(title="Biểu đồ Tròn")
    pie_sheet.add_image(img2, 'A1')
    print("Đã thêm thành công")

def save_plot_chart(visitor_data, workbook):
    city_methor_by_income = visitor_data.groupby("City")["Income"].sum()
    plt.figure(figsize=(13, 4))
    plt.plot(city_methor_by_income.index, city_methor_by_income.values)
    plt.savefig("City_by_income.png")
    plt.close()
    img3 = Image("City_by_income.png")
    pie_sheet = workbook.create_sheet(title="Biểu đồ đường")
    pie_sheet.add_image(img3, 'A1')
    print("Đã thêm thành công")

def main():
    """Hàm chính để thực hiện các bước."""
    file_path = "visitor_data_with_chart.xlsx"
    workbook = existfile(file_path)
    visitor_data = pd.read_excel("convert/Company data convert.xlsx")

    save_bar_chart(visitor_data, workbook)
    save_pie_chart(visitor_data, workbook)
    save_plot_chart(visitor_data, workbook)

    workbook.save(file_path) 

if __name__ == "__main__":
    main()
